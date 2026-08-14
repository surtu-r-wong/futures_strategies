from dataclasses import asdict

import pandas as pd
import pytest
from openpyxl import load_workbook

from common.metrics import summarize
from cta_carry.config import CarryConfig
from scripts.carry.compare_execution import (
    ComparisonInputError,
    compare_workbooks,
    main,
)


PERFORMANCE_METRICS = (
    "gross_ann_return",
    "net_ann_return",
    "gross_sharpe",
    "net_sharpe",
    "gross_ann_vol",
    "net_ann_vol",
    "gross_max_drawdown",
    "net_max_drawdown",
    "gross_calmar",
    "net_calmar",
    "annual_turnover",
    "total_cost",
    "annualized_cost",
    "avg_gross_leverage",
    "max_gross_leverage",
)


def _run_config(*, minute: bool, cost_bps: float = 4.0) -> pd.DataFrame:
    config = asdict(CarryConfig(cost_bps=cost_bps))
    rows = [
        {"key": "requested_start", "value": "2024-01-02"},
        {"key": "requested_end", "value": "2024-01-05"},
        {"key": "report_start_date", "value": "2024-01-02"},
        {"key": "products", "value": "A,B,C,D,E"},
        {"key": "code_diff_sha256", "value": ""},
        *({"key": key, "value": value} for key, value in config.items()),
        {"key": "execution_mode", "value": "minute" if minute else "daily"},
        {
            "key": "accounting_clock",
            "value": "piecewise_close_marked" if minute else "daily_open_marked",
        },
    ]
    if minute:
        rows.extend(
            [
                {"key": "session_rules_version", "value": "commodity-v1"},
                {"key": "minute_rows", "value": 1000},
            ]
        )
    return pd.DataFrame(rows)


def _daily_returns(*, minute: bool) -> pd.DataFrame:
    gross = [0.0, 0.012, -0.006, 0.008]
    turnover = [0.10, 0.20, 0.05, 0.15]
    cost = [0.00004, 0.00008, 0.00002, 0.00006]
    if minute:
        gross = [0.0, 0.010, -0.004, 0.011]
        turnover = [0.10, 0.25, 0.05, 0.20]
        cost = [0.00004, 0.00010, 0.00002, 0.00008]
    return pd.DataFrame(
        {
            "trade_date": pd.date_range("2024-01-02", periods=4, freq="D"),
            "gross_return": gross,
            "turnover": turnover,
            "cost": cost,
            "net_return": [value - fee for value, fee in zip(gross, cost, strict=True)],
            "gross_leverage": [0.0, 1.0, 1.5, 0.5],
        }
    )


def _executions() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "execution_id": "target-1",
                "trade_date": "2024-01-03",
                "product": "A",
                "contract": "A2405.SHF",
                "candidate_role": "signal_main",
                "execution_kind": "daily_target",
                "reason": "entry",
                "daily_open": 100.0,
                "vwap": 101.0,
                "volume": 10.0,
            },
            {
                "execution_id": "target-2",
                "trade_date": "2024-01-04",
                "product": "A",
                "contract": "A2405.SHF",
                "candidate_role": "carried",
                "execution_kind": "daily_target",
                "reason": "rebalance",
                "daily_open": 100.0,
                "vwap": 99.0,
                "volume": 30.0,
            },
            {
                "execution_id": "stop-1",
                "trade_date": "2024-01-04",
                "product": "A",
                "contract": "A2405.SHF",
                "candidate_role": "carried",
                "execution_kind": "intraday_stop",
                "reason": "stop_1",
                "daily_open": 100.0,
                "vwap": 120.0,
                "volume": 100.0,
            },
        ]
    )


def _intraday_stops() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "trade_date": "2024-01-04",
                "product": "A",
                "triggered": True,
                "execution_id": "stop-1",
            },
            {
                "trade_date": "2024-01-04",
                "product": "A",
                "triggered": True,
                "execution_id": "stop-2",
            },
            {
                "trade_date": "2024-01-04",
                "product": "B",
                "triggered": False,
                "execution_id": None,
            },
        ]
    )


def write_fixture_workbook(path, *, minute: bool):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        # Deliberately bogus: compare_workbooks must recompute from daily_returns.
        pd.DataFrame([{"ann_return": 999.0, "sharpe": 999.0}]).to_excel(
            writer,
            sheet_name="metrics",
            index=False,
        )
        _daily_returns(minute=minute).to_excel(
            writer,
            sheet_name="daily_returns",
            index=False,
        )
        if minute:
            _executions().to_excel(writer, sheet_name="executions", index=False)
            _intraday_stops().to_excel(
                writer,
                sheet_name="intraday_stops",
                index=False,
            )
        _run_config(minute=minute).to_excel(
            writer,
            sheet_name="run_config",
            index=False,
        )
    return path


def test_compare_execution_reports_minute_minus_daily(tmp_path):
    daily = write_fixture_workbook(tmp_path / "daily.xlsx", minute=False)
    minute = write_fixture_workbook(tmp_path / "minute.xlsx", minute=True)

    result = compare_workbooks(daily, minute, label="report_window")
    row = result.iloc[0]

    assert row["label"] == "report_window"
    for metric in PERFORMANCE_METRICS:
        assert row[f"{metric}_delta"] == pytest.approx(
            row[f"{metric}_minute"] - row[f"{metric}_daily"],
            nan_ok=True,
        )

    daily_frame = _daily_returns(minute=False)
    expected_net = summarize(
        daily_frame.set_index("trade_date")["net_return"],
        periods_per_year=252,
        turnover=daily_frame.set_index("trade_date")["turnover"],
    )
    assert row["net_ann_return_daily"] == pytest.approx(expected_net["ann_return"])
    assert row["net_ann_return_daily"] != 999.0
    assert row["stop_row_count"] == 3
    assert row["triggered_stop_count"] == 2
    assert row["same_day_multi_stop_count"] == 1
    assert row["daily_target_vwap_open_bps"] == pytest.approx(-50.0)
    assert row["gross_gap_explained_fraction"] == pytest.approx(
        row["gross_ann_return_delta"] / 0.10
    )

    expected_annual_turnover = daily_frame["turnover"].sum() * 252 / len(daily_frame)
    expected_total_cost = daily_frame["cost"].sum()
    assert row["annual_turnover_daily"] == pytest.approx(expected_annual_turnover)
    assert row["total_cost_daily"] == pytest.approx(expected_total_cost)
    assert row["annualized_cost_daily"] == pytest.approx(
        expected_total_cost * 252 / len(daily_frame)
    )


def _remove_sheet_column(path, sheet: str, column: str) -> None:
    workbook = load_workbook(path)
    worksheet = workbook[sheet]
    headers = [cell.value for cell in worksheet[1]]
    worksheet.delete_cols(headers.index(column) + 1)
    workbook.save(path)


def _replace_sheet_column(path, sheet: str, column: str, values) -> None:
    workbook = load_workbook(path)
    worksheet = workbook[sheet]
    headers = [cell.value for cell in worksheet[1]]
    column_number = headers.index(column) + 1
    for row_number, value in enumerate(values, start=2):
        worksheet.cell(row=row_number, column=column_number, value=value)
    workbook.save(path)


@pytest.mark.parametrize(
    ("sheet", "column"),
    [
        ("executions", "execution_kind"),
        ("intraday_stops", "triggered"),
    ],
)
def test_compare_execution_requires_the_final_audit_schema(
    tmp_path,
    sheet,
    column,
):
    daily = write_fixture_workbook(tmp_path / "daily.xlsx", minute=False)
    minute = write_fixture_workbook(tmp_path / "minute.xlsx", minute=True)
    _remove_sheet_column(minute, sheet, column)

    with pytest.raises(ComparisonInputError) as exc_info:
        compare_workbooks(daily, minute, label="report_window")

    error = exc_info.value
    assert error.path == minute
    assert error.check == "column_missing"
    assert error.sheet == sheet
    assert column in error.reason


def test_compare_execution_accepts_empty_intraday_stops_with_schema(tmp_path):
    daily = write_fixture_workbook(tmp_path / "daily.xlsx", minute=False)
    minute = write_fixture_workbook(tmp_path / "minute.xlsx", minute=True)
    workbook = load_workbook(minute)
    worksheet = workbook["intraday_stops"]
    worksheet.delete_rows(2, worksheet.max_row)
    workbook.save(minute)

    result = compare_workbooks(daily, minute, label="report_window")
    row = result.iloc[0]

    assert row["stop_row_count"] == 0
    assert row["triggered_stop_count"] == 0
    assert row["same_day_multi_stop_count"] == 0


def test_compare_execution_still_rejects_other_empty_required_sheets(tmp_path):
    daily = write_fixture_workbook(tmp_path / "daily.xlsx", minute=False)
    minute = write_fixture_workbook(tmp_path / "minute.xlsx", minute=True)
    workbook = load_workbook(minute)
    worksheet = workbook["executions"]
    worksheet.delete_rows(2, worksheet.max_row)
    workbook.save(minute)

    with pytest.raises(ComparisonInputError) as exc_info:
        compare_workbooks(daily, minute, label="report_window")

    error = exc_info.value
    assert error.path == minute
    assert error.check == "sheet_empty"
    assert error.sheet == "executions"


def test_compare_execution_rejects_missing_required_sheet_structurally(tmp_path):
    daily = write_fixture_workbook(tmp_path / "daily.xlsx", minute=False)
    minute = write_fixture_workbook(tmp_path / "minute.xlsx", minute=True)
    workbook = load_workbook(minute)
    del workbook["executions"]
    workbook.save(minute)

    with pytest.raises(ComparisonInputError) as exc_info:
        compare_workbooks(daily, minute, label="report_window")

    error = exc_info.value
    assert error.path == minute
    assert error.check == "sheet_missing"
    assert error.sheet == "executions"


def test_compare_execution_rejects_research_parameter_mismatch(tmp_path):
    daily = write_fixture_workbook(tmp_path / "daily.xlsx", minute=False)
    minute = write_fixture_workbook(
        tmp_path / "minute.xlsx",
        minute=True,
    )
    workbook = load_workbook(minute)
    worksheet = workbook["run_config"]
    keys = [cell.value for cell in worksheet["A"]]
    row_number = keys.index("cost_bps") + 1
    worksheet.cell(row=row_number, column=2, value=1.3)
    workbook.save(minute)

    with pytest.raises(ComparisonInputError) as exc_info:
        compare_workbooks(daily, minute, label="report_window")

    error = exc_info.value
    assert error.check == "paired_config_mismatch"
    assert error.key == "cost_bps"


def test_compare_execution_rejects_different_report_trading_dates(tmp_path):
    daily = write_fixture_workbook(tmp_path / "daily.xlsx", minute=False)
    minute = write_fixture_workbook(tmp_path / "minute.xlsx", minute=True)
    workbook = load_workbook(minute)
    worksheet = workbook["daily_returns"]
    worksheet.cell(row=worksheet.max_row, column=1, value="2024-01-08")
    workbook.save(minute)

    with pytest.raises(ComparisonInputError) as exc_info:
        compare_workbooks(daily, minute, label="report_window")

    error = exc_info.value
    assert error.check == "paired_report_dates"
    assert error.sheet == "daily_returns"
    assert error.key == "trade_date"


@pytest.mark.parametrize(
    ("return_column", "returns"),
    [
        ("gross_return", [0.0, -1.0, 0.02, 0.01]),
        ("net_return", [0.0, -1.5, -1.5, 0.01]),
    ],
)
def test_compare_execution_rejects_any_depleted_equity_path(
    tmp_path,
    return_column,
    returns,
):
    daily = write_fixture_workbook(tmp_path / "daily.xlsx", minute=False)
    minute = write_fixture_workbook(tmp_path / "minute.xlsx", minute=True)
    _replace_sheet_column(
        daily,
        "daily_returns",
        return_column,
        returns,
    )

    with pytest.raises(ComparisonInputError) as exc_info:
        compare_workbooks(daily, minute, label="report_window")

    error = exc_info.value
    assert error.path == daily
    assert error.check == "equity_depleted"
    assert error.sheet == "daily_returns"
    assert error.key == return_column


def _replace_with_zero_return_series(path) -> None:
    _replace_sheet_column(
        path,
        "daily_returns",
        "gross_return",
        [0.0, 0.0, 0.0, 0.0],
    )
    _replace_sheet_column(
        path,
        "daily_returns",
        "net_return",
        [0.0, 0.0, 0.0, 0.0],
    )


def test_compare_execution_rejects_nonfinite_recomputed_metric(tmp_path):
    daily = write_fixture_workbook(tmp_path / "daily.xlsx", minute=False)
    minute = write_fixture_workbook(tmp_path / "minute.xlsx", minute=True)
    _replace_with_zero_return_series(daily)

    with pytest.raises(ComparisonInputError) as exc_info:
        compare_workbooks(daily, minute, label="report_window")

    error = exc_info.value
    assert error.path == daily
    assert error.check == "performance_metric_nonfinite"
    assert error.sheet == "daily_returns"
    assert error.key == "gross_sharpe"


def test_compare_execution_cli_rejects_nonfinite_metric_without_output(
    tmp_path,
    capsys,
):
    daily = write_fixture_workbook(tmp_path / "daily.xlsx", minute=False)
    minute = write_fixture_workbook(tmp_path / "minute.xlsx", minute=True)
    _replace_with_zero_return_series(daily)
    output = tmp_path / "comparison.csv"

    code = main(
        [
            "--pair",
            "report_window",
            str(daily),
            str(minute),
            "--output",
            str(output),
        ]
    )

    assert code == 2
    assert not output.exists()
    error = capsys.readouterr().err
    assert "check=performance_metric_nonfinite" in error
    assert "key=gross_sharpe" in error


def test_compare_execution_cli_writes_repeated_pairs(tmp_path):
    daily = write_fixture_workbook(tmp_path / "daily.xlsx", minute=False)
    minute = write_fixture_workbook(tmp_path / "minute.xlsx", minute=True)
    output = tmp_path / "nested" / "comparison.csv"

    code = main(
        [
            "--pair",
            "report_window",
            str(daily),
            str(minute),
            "--pair",
            "full_history",
            str(daily),
            str(minute),
            "--output",
            str(output),
        ]
    )

    assert code == 0
    written = pd.read_csv(output)
    assert written["label"].tolist() == ["report_window", "full_history"]
    assert written.columns.tolist() == [
        "label",
        *(
            f"{metric}_{suffix}"
            for metric in PERFORMANCE_METRICS
            for suffix in ("daily", "minute", "delta")
        ),
        "stop_row_count",
        "triggered_stop_count",
        "same_day_multi_stop_count",
        "daily_target_vwap_open_bps",
        "gross_gap_explained_fraction",
    ]
